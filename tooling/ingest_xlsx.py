#!/usr/bin/env python3
"""Convert the approved NEDS Q&A spreadsheet into data/faq.json.

This is the Python side of the split: an OFFLINE tool that turns the source of
truth (the "Q&A Bank" sheet of the NEDS chatbot question bank xlsx) into the
JSON the TypeScript service reads. It does not run in the request path and
shares no runtime with the service.

Usage:
    python3 tooling/ingest_xlsx.py path/to/NEDS-chatbot-question-bank.xlsx

Reads the "Q&A Bank" sheet (falls back to the active sheet), auto-detects the
header row (the one starting with "ID"), and keeps a row's "Draft chatbot
answer" as trusted when EITHER:
  - "Office-approved answer" is "PK Reviewed" or "Ad Reviewed" (office sign-off), or
  - "Confidence" is "High" or "Medium" (confident research, even if not yet
    individually signed off).
Rows marked "DELETE" or "Not Relevant" in "Office-approved answer" are always
dropped, even if their confidence is high — an explicit reviewer decision
overrides the confidence score. Everything else (blank/low/not-assessed
confidence and not office-reviewed) is left out as not yet trustworthy.

Writes data/faq.json in the shape the service expects. Restart the service
after to pick up the new knowledge.
"""

from __future__ import annotations

import json
import re
import sys
from pathlib import Path

try:
    from openpyxl import load_workbook
except ImportError:
    sys.exit(
        "openpyxl is required. Create a venv and install it:\n"
        "  python3 -m venv tooling/.venv\n"
        "  tooling/.venv/bin/pip install -r tooling/requirements.txt"
    )

ROOT = Path(__file__).resolve().parent.parent
OUT = ROOT / "data" / "faq.json"

APPROVED_STATUSES = {"pk reviewed", "ad reviewed"}
EXCLUDED_STATUSES = {"delete", "not relevant"}
TRUSTED_CONFIDENCE = {"high", "medium"}


def slug(text: str) -> str:
    s = re.sub(r"[^a-z0-9]+", "-", text.lower()).strip("-")
    return s[:48] or "faq"


def norm(header: str) -> str:
    return re.sub(r"[^a-z]", "", (header or "").lower())


def find_header_row(rows: list[tuple]) -> int:
    for i, row in enumerate(rows):
        if row and norm(str(row[0])) == "id":
            return i
    sys.exit("Could not find the header row (expected a row starting with 'ID').")


def main(xlsx_path: str) -> None:
    wb = load_workbook(xlsx_path, read_only=True, data_only=True)
    ws = wb["Q&A Bank"] if "Q&A Bank" in wb.sheetnames else wb.active
    rows = list(ws.iter_rows(values_only=True))

    header_i = find_header_row(rows)
    header = [norm(str(c)) for c in rows[header_i]]

    def col(name: str) -> int:
        if name not in header:
            sys.exit(f"Spreadsheet must have a '{name}' column.")
        return header.index(name)

    q_i = col("customerquestion")
    a_i = col("draftchatbotanswer")
    cat_i = col("category")
    status_i = col("officeapprovedanswer")
    conf_i = col("confidence")
    id_i = col("id")

    faqs = []
    seen: set[str] = set()
    skipped_excluded = skipped_unreviewed = skipped_empty = 0
    for row in rows[header_i + 1 :]:
        if not row or len(row) <= max(q_i, a_i, status_i, conf_i):
            continue
        status = str(row[status_i] or "").strip().lower()
        confidence = str(row[conf_i] or "").strip().lower()
        if status in EXCLUDED_STATUSES:
            skipped_excluded += 1
            continue
        if status not in APPROVED_STATUSES and confidence not in TRUSTED_CONFIDENCE:
            skipped_unreviewed += 1
            continue
        q = str(row[q_i] or "").strip()
        a = str(row[a_i] or "").strip()
        if not q or not a or a.lower() == "none":
            skipped_empty += 1
            continue

        raw_id = row[id_i]
        fid = str(int(raw_id)) if isinstance(raw_id, float) else (str(raw_id) if raw_id else slug(q))
        while fid in seen:
            fid += "-x"
        seen.add(fid)

        category = str(row[cat_i] or "").strip()
        keywords = [w.lower() for w in re.split(r"[^A-Za-z0-9+]+", category) if w]

        faqs.append({"id": fid, "question": q, "answer": a, "keywords": keywords})

    payload = {
        "_note": (
            f"Generated from {Path(xlsx_path).name} ('Q&A Bank' sheet) by "
            "tooling/ingest_xlsx.py. Includes rows office-approved as 'PK "
            "Reviewed'/'Ad Reviewed', or with 'High'/'Medium' research "
            "confidence; always excludes 'DELETE'/'Not Relevant' rows even if "
            "confident. Do not edit by hand — re-run the ingest script instead."
        ),
        "faqs": faqs,
    }
    OUT.write_text(json.dumps(payload, indent=2, ensure_ascii=False), encoding="utf-8")
    print(f"Wrote {len(faqs)} approved Q&A entries to {OUT.relative_to(ROOT)}")
    print(
        f"Skipped: {skipped_excluded} delete/not-relevant, "
        f"{skipped_unreviewed} not reviewed and not high/medium confidence, "
        f"{skipped_empty} empty/blank"
    )


if __name__ == "__main__":
    if len(sys.argv) != 2:
        sys.exit("Usage: python3 tooling/ingest_xlsx.py path/to/neds-qa.xlsx")
    main(sys.argv[1])
